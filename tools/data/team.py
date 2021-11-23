# -*- coding: utf-8 -*-

import sys
import os
import re
import json
import django
from django.core.files.base import ContentFile, File
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

MY_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.abspath(MY_DIR+'/../../')
sys.path.append(BASE_DIR)
os.environ['DJANGO_SETTINGS_MODULE'] = 'config.settings_local'
django.setup()

from front import models
from control.serializers.common import CommonTeamPeopleSerializer


def parse():

    db_locations = dict([(
        location.title.lower().split(' (')[0], location.id
    ) for location in models.CommonLocation.objects.all()])

    new_locations = dict()

    print db_locations

    db_positions = dict([(
        position.title.lower(), position.id
    ) for position in models.CommonTeamPeoplePosition.objects.all()])

    new_positions = dict()

    print db_positions

    # return

    db_teams = dict([(
        team[1].split(' ')[0].lower(), team[0]
    ) for team in models.CommonTeamPeople.TeamChoices])

    print db_teams

    teams_mapping = {
        'synthesis': 'development',
        'thrive': 'technologies'
    }

    # return

    print "Parse..."

    workbook = load_workbook(os.path.join(MY_DIR, 'team.xlsx'))
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    print sheet_name

    members = []

    row_num = 2

    _team = None

    email_re = re.compile(r'.*\"(.+)\".*\"(.+)\"')

    # models.CommonTeamPeople.objects.all().delete()

    for row in sheet.iter_rows(min_row=2, max_col=15):
        first_name = row[0].value

        if not first_name:
            break

        first_name = first_name.strip()

        id = row[14].value

        if not id:
            id = None

        last_name = row[1].value.strip()
        team = row[2].value

        if team:
            _team = team
        else:
            team = _team

        team = str(team).lower()

        team = db_teams.get(teams_mapping.get(team, team), None)

        title = row[3].value

        if title:
            title = title.strip()

        location = row[4].value

        if location:
            location = location.strip().lower()

        location_id = db_locations.get(location, None)

        if location and not location_id:

            if location not in new_locations:
                new_locations[location] = None

            if not location_id:
                location_id = (location, location_id)

        leadership = row[5].value

        leadership = True if leadership and leadership.strip().lower() == 'yes' else False

        positions = []

        position = row[6].value

        if position:
            position = position.strip().lower()
            position_id = db_positions.get(position, None)

            if position not in new_positions:
                new_positions[position] = None

            if not position_id:
                position_id = (position, position_id)

            positions.append(position_id)

        position = row[7].value
        if position:
            position = position.strip().lower()
            position_id = db_positions.get(position, None)

            if position not in new_positions:
                new_positions[position] = None

            if not position_id:
                position_id = (position, position_id)

            positions.append(position_id)

        email = row[12].value

        if email:

            email_res = email_re.search(email)
            if email_res:
                email = email_res.group(2)

            email = email.strip()

        square = None
        rectangle = None

        photos = row[8].value
        if photos:

            if photos.find('Headshot') != -1:

                path = os.path.join(MY_DIR, 'team', 'SqBW_%s %s.jpg' % (
                    first_name, last_name
                ))

                if os.path.isfile(path):
                    square = path

                path = os.path.join(MY_DIR, 'team', 'SqBW_%s %s.jpg' % (
                    first_name.replace(' ', '-'), last_name.replace(' ', '-')
                ))

                if os.path.isfile(path):
                    square = path

                path = os.path.join(MY_DIR, 'team', 'SqBW_%s %s.png' % (
                    first_name, last_name
                ))

                if os.path.isfile(path):
                    square = path

                path = os.path.join(MY_DIR, 'team', 'SqBW_%s %s.png' % (
                    first_name.replace(' ', '-'), last_name.replace(' ', '-')
                ))

                if os.path.isfile(path):
                    square = path

            if photos.find('Casual') != -1:

                path = os.path.join(MY_DIR, 'team', 'Candid_%s %s.jpg' % (
                    first_name, last_name
                ))

                if os.path.isfile(path):
                    rectangle = path

                path = os.path.join(MY_DIR, 'team', 'Candid_%s %s.jpg' % (
                    first_name.replace(' ', '-'), last_name.replace(' ', '-')
                ))

                if os.path.isfile(path):
                    rectangle = path

                path = os.path.join(MY_DIR, 'team', 'Candid_%s %s.png' % (
                    first_name, last_name
                ))

                if os.path.isfile(path):
                    rectangle = path

                path = os.path.join(MY_DIR, 'team', 'Candid_%s %s.png' % (
                    first_name.replace(' ', '-'), last_name.replace(' ', '-')
                ))

                if os.path.isfile(path):
                    rectangle = path

        # print id, first_name, last_name, team, title, location_id, leadership, positions, email

        # member, created = models.CommonTeamPeople.objects.get_or_create(id=id)

        if isinstance(location_id, tuple):
            print "Non exist location", location_id

        member_positions = list()

        for position in positions:

            if isinstance(position, tuple):

                if position[0] not in new_positions or not new_positions[position[0]]:

                    print "Non exist position", position
                    print "Creating location..."

                    db_position = models.CommonTeamPeoplePosition.objects.create(title=position[0])

                    new_positions[position[0]] = db_position.id

                member_positions.append({
                    'position_id': new_positions[position[0]]
                })
            else:
                member_positions.append({
                    'position_id': position
                })

        if square:
            square = File(open(square, 'rb'), name=os.path.basename(square))

        if rectangle:
            rectangle = File(open(rectangle, 'rb'), name=os.path.basename(rectangle))

        member = {
            'id': id,
            'name': ('%s %s' % (first_name, last_name)).strip(),
            'team': team,
            'title': title,
            'location': None if isinstance(location_id, tuple) else location_id,
            'leadership': leadership,
            'positions': member_positions,
            'email': email,
            'image': square,
            'image2': rectangle,
            'order': row[0].row - 2
        }

        print member

        try:
            db_member = models.CommonTeamPeople.objects.get(id=id) if id else None
        except models.CommonTeamPeople.DoesNotExist:
            db_member = None
            id = None

        serializer = CommonTeamPeopleSerializer(
            data=member,
            instance=db_member
        )

        if not serializer.is_valid():
            print "Invalid data:"
            print serializer.errors
            break

        member_obj = serializer.save()

        sheet.cell(row=row[0].row, column=15).value = member_obj.id

        # if row[0].row > 2:
        #     break

    # print json.dumps(members, indent=4)

    # print json.dumps(new_locations.keys())

    workbook.save(os.path.join(MY_DIR, 'team.xlsx'))


if __name__ == '__main__':
    parse()
